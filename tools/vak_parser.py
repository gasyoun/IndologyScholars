"""Parser for the VAK (Higher Attestation Commission) peer-reviewed journal list.

Supports both Excel (.xlsx) and PDF formats. Auto-detects format from extension.

PDF source: https://cpr.bmstu.ru/sites/default/files/.../Перечень журналов ВАК 29.04.2026.pdf
Excel source: https://vak.minobrnauki.gov.ru/documents#tab=_tab:ref~

Usage:
  python tools/vak_parser.py path/to/perechen_vak.xlsx
  python tools/vak_parser.py path/to/perechen_vak.pdf
  python tools/vak_parser.py --pdf html_cache/vak/perechen_2026.pdf

Outputs:
  analytics_output/vak_journals.csv               — all journals
  analytics_output/vak_journals_philology.csv      — philology only (5.9.x / 10.x)
  editors/<journal_slug>.md                        — one profile per journal

Requirements: pip install openpyxl pymupdf
"""

import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EDITORS_DIR = ROOT / "editors"
ANALYTICS_OUT = ROOT / "analytics_output"

PHILOLOGY_CODES = re.compile(r"5\.9\.\d+|10\.01\.\d+|10\.02\.\d+")

INDOLOGY_KEYWORDS = [
    "восток", "oriental", "india", "инди", "азия", "asia",
    "восточн", "письменн", "памятник", "санскрит", "sanskrit",
    "филолог", "philolog", "лингвист", "linguist", "тибет",
    "будд", "buddh", "индолог",
]


def normalize_slug(name: str) -> str:
    slug = name.lower().strip()
    slug = re.sub(r"[^a-zа-яё0-9]+", "-", slug)
    return slug.strip("-")[:80]


# ── PDF PARSER ────────────────────────────────────────────────────────

def parse_vak_pdf(pdf_path: str) -> list[dict]:
    """Parse the VAK PDF journal list into a list of journal dicts.

    The PDF table has columns: №, Наименование, ISSN, Специальности, Дата.
    Each journal entry spans multiple text lines.
    ISSN is the reliable anchor for splitting entries.
    """
    try:
        import fitz
    except ImportError:
        sys.exit("Install pymupdf: pip install pymupdf")

    doc = fitz.open(pdf_path)
    all_text = ""
    for page in doc:
        all_text += page.get_text("text") + "\n"
    doc.close()

    # Split into entries using ISSN as anchor
    # ISSN pattern: 4 digits - 4 digits or digits-X
    issn_re = re.compile(r"(\d{4}-\d{3}[\dXx])")

    # Find all ISSN positions
    issn_matches = list(issn_re.finditer(all_text))
    print(f"  Found {len(issn_matches)} ISSN entries in PDF")

    journals = []
    for i, m in enumerate(issn_matches):
        issn = m.group(1)
        issn_pos = m.start()

        # Get text window around ISSN: from previous ISSN to next ISSN
        prev_end = issn_matches[i - 1].start() if i > 0 else 0
        next_start = issn_matches[i + 1].start() if i + 1 < len(issn_matches) else len(all_text)
        window = all_text[prev_end:next_start]

        # Find the entry number (digits followed by period) before ISSN
        pre_issn = window[:window.index(issn)].strip()
        pre_lines = pre_issn.split("\n")
        num_line_idx = -1
        for idx in range(len(pre_lines) - 1, -1, -1):
            if re.match(r"^\d{1,4}\.\s*$", pre_lines[idx].strip()):
                num_line_idx = idx
                break

        # Title: lines between entry number and ISSN (end of pre_issn)
        if num_line_idx >= 0:
            title_lines = pre_lines[num_line_idx + 1:]
        else:
            # Fallback: take the few lines right before ISSN
            title_lines = [l.strip() for l in pre_lines[-8:] if l.strip() and len(l.strip()) > 3]

        title_parts = []
        for line in title_lines:
            line = line.strip()
            if not line or len(line) < 4:
                continue
            # Skip date lines
            if re.match(r"^[сc]\s+\d{2}\.\d{2}\.\d{4}", line):
                break
            # Skip pure code lines
            if re.match(r"^(?:\d{2}\.\d{2}\.\d{2}[.,\s;]*)+$", line):
                break
            # Skip "по ДД.ММ.ГГГГ" lines
            if re.match(r"^по\s+\d{2}\.\d{2}\.\d{4}", line):
                break
            # Skip lines that look like ISSN already
            if re.match(r"^\d{4}-\d{3}[\dXx]$", line.strip()):
                break
            title_parts.append(line)

        title = " ".join(title_parts)
        title = re.sub(r"\s+", " ", title).strip()
        # Clean up: remove trailing ISSN if accidentally included
        title = re.sub(r"\s*\d{4}-\d{3}[\dXx]\s*", "", title)

        if not title or len(title) < 10:
            continue

        # Specialty codes: text after ISSN
        post_issn = window[window.index(issn) + len(issn):].strip()
        codes = []
        for m2 in re.finditer(r"(\d{1,2}\.\d{1,2}\.\d{1,2})", post_issn):
            codes.append(m2.group(1))
        specialty = "; ".join(codes[:20])  # max 20 codes

        # Date: find "с ДД.ММ.ГГГГ" or "c ДД.ММ.ГГГГ"
        date_match = re.search(r"[сc]\s+(\d{2}\.\d{2}\.\d{4})", post_issn)
        date_included = date_match.group(1) if date_match else ""

        if title and issn:
            # Clean up title: remove registration notes, old name references
            title = re.sub(r"\s*\(\s*[сc]\s+\d{2}\.\d{2}\.\d{4}.*?\)", "", title)
            title = re.sub(r"\s*\(\s*ранее\s+.*?\)", "", title)
            title = re.sub(r"\s*\(\s*прежнее\s+.*?\)", "", title)
            title = re.sub(r"\s+", " ", title).strip()
            if not title or len(title) < 5:
                continue
            journals.append({
                "title": title,
                "issn": issn,
                "specialty": specialty,
                "date_included": date_included,
            })

    return journals


def is_philology(specialty: str) -> bool:
    return bool(PHILOLOGY_CODES.search(specialty))


def is_indology_relevant(title: str, specialty: str) -> bool:
    combined = f"{title} {specialty}".lower()
    for kw in INDOLOGY_KEYWORDS:
        if kw in combined:
            return True
    return False


def generate_editor_profile(journal: dict) -> str:
    slug = normalize_slug(journal["title"])
    title = journal["title"]
    issn = journal.get("issn", "")
    specialty = journal.get("specialty", "")
    return f"""# {title}

- **ISSN:** {issn}
- **Специальности ВАК:** {specialty}
- **Профиль:** [добавить ссылку на сайт журнала]
- **Статус в перечне ВАК:** включён

## Редакционная коллегия

<!-- Заполнить вручную -->

## Профиль публикаций

<!-- Типичные темы, языки, объём статей -->

## Заметки для подачи

<!-- Особые требования, опыт подачи -->

## Контакты

<!-- E-mail редакции, адрес для отправки -->
"""


def load_journals(path: str) -> list[dict]:
    """Auto-detect format and parse."""
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return parse_vak_pdf(path)
    else:
        # Try Excel
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        journals = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            row_values = [str(c).strip() if c else "" for c in row]
            if not any(row_values):
                continue
            if headers is None:
                if any("issn" in c.lower() or "назван" in c.lower() for c in row_values):
                    headers = row_values
                    continue
                headers = [f"col_{i}" for i in range(len(row_values))]
                if any("issn" in c.lower() for c in row_values):
                    continue
            else:
                j = {}
                for i, h in enumerate(headers):
                    j[h] = row_values[i] if i < len(row_values) else ""
                title = j.get("title") or j.get("название") or j.get("col_1", "")
                issn = j.get("issn") or j.get("col_2", "")
                specialty = j.get("specialty") or j.get("специальность") or j.get("col_3", "")
                if title and len(title) >= 3:
                    journals.append({"title": title, "issn": issn, "specialty": specialty, "date_included": ""})
        wb.close()
        return journals


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = sys.argv[-1]
    if not Path(filepath).exists():
        sys.exit(f"File not found: {filepath}")

    print(f"Parsing {filepath}...")
    journals = load_journals(filepath)

    if not journals:
        print("No journals parsed. Check file format.")
        sys.exit(1)

    print(f"Parsed {len(journals)} journals total")

    phil_journals = [j for j in journals if is_philology(j["specialty"])]
    print(f"Philology journals (5.9.x / 10.01.x): {len(phil_journals)}")

    indo_journals = [j for j in phil_journals if is_indology_relevant(j["title"], j["specialty"])]
    print(f"Indology/Oriental-relevant: {len(indo_journals)}")
    for j in indo_journals[:20]:
        print(f"  - {j['title'][:60]} [{j.get('issn', '')}]")

    ANALYTICS_OUT.mkdir(parents=True, exist_ok=True)

    fields = ["title", "issn", "specialty"]
    with open(ANALYTICS_OUT / "vak_journals.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for j in journals:
            w.writerow({k: j.get(k, "") for k in fields})

    EDITORS_DIR.mkdir(parents=True, exist_ok=True)
    profile_count = 0
    indo_count = 0
    for j in phil_journals:
        slug = normalize_slug(j["title"])
        (EDITORS_DIR / f"{slug}.md").write_text(generate_editor_profile(j), encoding="utf-8")
        profile_count += 1
        if is_indology_relevant(j["title"], j["specialty"]):
            indo_count += 1

    # Write Indology-specific CSV
    with open(ANALYTICS_OUT / "vak_journals_indology.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for j in phil_journals:
            if is_indology_relevant(j["title"], j["specialty"]):
                w.writerow({k: j.get(k, "") for k in fields})

    print(f"\nOutputs: vak_journals.csv ({len(journals)} journals), "
          f"vak_journals_indology.csv ({indo_count} indology), editors/*.md ({profile_count} profiles)")
    print(f"\nPhilology journals:")
    for j in phil_journals:
        print(f"  {j['title'][:70]:70s}  ISSN: {j.get('issn', 'N/A')}")


if __name__ == "__main__":
    main()
