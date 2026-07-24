"""Ingest the VK wall export ``vk_posts_all.xlsx`` into a searchable SQLite database.

The export (one row per post: id, permalink, date, text, likes, reposts,
comments, views, attachment count, attachment types) covers the public VK
wall behind "Общество ревнителей санскрита" (vk.com/wall-88831040), 2015 to
present. Unlike the closed nagari@googlegroups.com Takeout, there is no
per-author or thread structure — one publishing account, no reply graph — so
the schema is flatter than ``nagari_group_archive.ingest``. Writes:

* ``posts``      — one row per post, with engagement counters and parsed
  attachment types.
* ``posts_fts``  — an FTS5 index over the post text for ranked full-text
  search across Russian, IAST and Devanagari (unicode61, diacritics folded
  so ``atman`` matches ``ātman``).
* ``attachments`` — optional wave-1a table loaded from
  ``data/attachments_raw.json`` (produced by ``fetch.py``). Additive; the
  xlsx remains read-only and its column layout is unchanged.

Usage::

    python -m vk_ors_archive.ingest --limit 300     # quick validation slice
    python -m vk_ors_archive.ingest                 # full run

Nothing here is destructive to the source: the xlsx is opened read-only.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl

PKG = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PKG / "vk_posts_all.xlsx"
DEFAULT_DB = PKG / "data" / "vk_ors.db"
DEFAULT_ATTACHMENTS = PKG / "data" / "attachments_raw.json"

RE_WS = re.compile(r"\s+")
RE_HASHTAG = re.compile(r"#([\w][\w\-]{1,60})", re.UNICODE)


def configure_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")


def parse_date(value) -> tuple[str, int | None, int | None, str]:
    """Return (iso, year, month, ym) from an openpyxl cell value."""
    dt = None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(value.strip(), fmt)
                break
            except ValueError:
                continue
    if dt is None:
        return "", None, None, ""
    return dt.isoformat(), dt.year, dt.month, f"{dt.year:04d}-{dt.month:02d}"


def parse_post_id(url: str, fallback: int) -> int:
    m = re.search(r"_(\d+)$", url or "")
    return int(m.group(1)) if m else fallback


SCHEMA = """
CREATE TABLE posts (
    id                INTEGER PRIMARY KEY,
    url               TEXT,
    date_utc          TEXT,
    year              INTEGER,
    month             INTEGER,
    ym                TEXT,
    text              TEXT,
    text_chars        INTEGER,
    likes             INTEGER,
    reposts           INTEGER,
    comments          INTEGER,
    views             INTEGER,
    n_attachments     INTEGER,
    attachment_types  TEXT
);
CREATE TABLE hashtags (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    post_id      INTEGER REFERENCES posts(id),
    tag          TEXT
);
CREATE TABLE attachments (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    post_id      INTEGER REFERENCES posts(id),
    type         TEXT,
    url          TEXT,
    width        INTEGER,
    height       INTEGER,
    position     INTEGER
);
CREATE INDEX idx_posts_ym    ON posts(ym);
CREATE INDEX idx_posts_year  ON posts(year);
CREATE INDEX idx_hashtags_tag ON hashtags(tag);
CREATE INDEX idx_hashtags_post ON hashtags(post_id);
CREATE INDEX idx_attachments_post ON attachments(post_id);
CREATE INDEX idx_attachments_type ON attachments(type);
CREATE VIRTUAL TABLE posts_fts USING fts5(
    text,
    tokenize = 'unicode61 remove_diacritics 2'
);
"""


def iter_rows(xlsx_path: Path, limit: int | None = None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Sheet1"]
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        yield row
        n += 1
        if limit and n >= limit:
            return


def parse_row(row: tuple, fallback_idx: int) -> dict:
    raw_id, url, date_raw, text, likes, reposts, comments, views, n_att, att_types = (
        list(row) + [None] * (10 - len(row))
    )[:10]
    iso, year, month, ym = parse_date(date_raw)
    post_id = int(raw_id) if raw_id is not None else parse_post_id(url, fallback_idx)
    text = (text or "").strip()
    tags = sorted({t.lower() for t in RE_HASHTAG.findall(text)})
    return {
        "id": post_id,
        "url": url or "",
        "date_utc": iso,
        "year": year,
        "month": month,
        "ym": ym,
        "text": text,
        "text_chars": len(text),
        "likes": int(likes or 0),
        "reposts": int(reposts or 0),
        "comments": int(comments or 0),
        "views": int(views or 0),
        "n_attachments": int(n_att or 0),
        "attachment_types": (att_types or "").strip(),
        "tags": tags,
    }


def load_attachments(db: sqlite3.Connection, path: Path, known_ids: set[int]) -> int:
    """Load attachments_raw.json into the attachments table. Returns row count."""
    if not path.exists():
        print(f"  attachments: {path} missing — skipping (run fetch.py first for gallery)", flush=True)
        return 0
    raw = json.loads(path.read_text(encoding="utf-8"))
    n = 0
    skipped_orphan = 0
    for post_id_s, items in raw.items():
        try:
            post_id = int(post_id_s)
        except (TypeError, ValueError):
            continue
        if post_id not in known_ids:
            # --limit slice: raw file has full wall; only load posts we ingested
            skipped_orphan += 1
            continue
        for item in items or []:
            db.execute(
                """INSERT INTO attachments(post_id, type, url, width, height, position)
                   VALUES (?,?,?,?,?,?)""",
                (
                    post_id,
                    (item.get("type") or "unknown"),
                    (item.get("url") or ""),
                    item.get("width"),
                    item.get("height"),
                    int(item.get("position") or 0),
                ),
            )
            n += 1
    if skipped_orphan:
        print(f"  attachments: skipped {skipped_orphan} post_ids not in posts (limit slice?)", flush=True)
    return n


def build(
    xlsx_path: Path,
    db_path: Path,
    limit: int | None,
    attachments_path: Path,
) -> dict:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"missing {xlsx_path}")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    db = sqlite3.connect(db_path)
    db.executescript(SCHEMA)

    n_posts = n_tags = 0
    known_ids: set[int] = set()
    start = time.time()
    for idx, row in enumerate(iter_rows(xlsx_path, limit=limit), start=1):
        rec = parse_row(row, idx)
        db.execute(
            """INSERT OR REPLACE INTO posts(id,url,date_utc,year,month,ym,text,text_chars,
               likes,reposts,comments,views,n_attachments,attachment_types)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                rec["id"], rec["url"], rec["date_utc"], rec["year"], rec["month"], rec["ym"],
                rec["text"], rec["text_chars"], rec["likes"], rec["reposts"], rec["comments"],
                rec["views"], rec["n_attachments"], rec["attachment_types"],
            ),
        )
        known_ids.add(rec["id"])
        for tag in rec["tags"]:
            db.execute("INSERT INTO hashtags(post_id,tag) VALUES (?,?)", (rec["id"], tag))
            n_tags += 1
        n_posts += 1
        if n_posts % 1000 == 0:
            db.commit()
            print(f"  ... {n_posts} posts, {time.time()-start:.0f}s", flush=True)

    print("  populating FTS index ...", flush=True)
    db.execute("INSERT INTO posts_fts(rowid,text) SELECT id,text FROM posts")

    print("  loading attachments ...", flush=True)
    n_att = load_attachments(db, attachments_path, known_ids)

    db.commit()
    db.execute("PRAGMA optimize")
    db.close()
    return {
        "posts": n_posts,
        "hashtags": n_tags,
        "attachments": n_att,
        "seconds": round(time.time() - start, 1),
    }


def main(argv: list[str] | None = None) -> None:
    configure_stdio()
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="VK export xlsx path")
    ap.add_argument("--db", type=Path, default=DEFAULT_DB, help="output SQLite path")
    ap.add_argument("--attachments", type=Path, default=DEFAULT_ATTACHMENTS,
                    help="attachments_raw.json from fetch.py")
    ap.add_argument("--limit", type=int, default=None, help="parse only the first N rows (validation)")
    args = ap.parse_args(argv)
    print(f"Ingesting {args.xlsx}", flush=True)
    stats = build(args.xlsx, args.db, args.limit, args.attachments)
    print(f"Done: {stats}", flush=True)
    print(f"DB: {args.db}", flush=True)


if __name__ == "__main__":
    main()
